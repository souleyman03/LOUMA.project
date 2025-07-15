import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tempfile

# Titre de l'application
st.title("📦 Générateur de Reporting Ventes SIM")

# Uploader du fichier Excel brut
uploaded_file = st.file_uploader("📁 Importer le fichier Excel brut (hebdomadaire)", type=["xlsx", "csv"])

if uploaded_file: 

    if uploaded_file.name.endswith('.csv'):
        df = pd.read_csv(uploaded_file, encoding='utf-8', sep=';')


    else:
        # Charger toutes les feuilles sans les lire entièrement
        xls = pd.ExcelFile(uploaded_file)
            
        # Afficher les noms de feuilles disponibles
        sheet_names = xls.sheet_names
        selected_sheet = st.selectbox("🗂️ Choisir la feuille à exploiter :", options=sheet_names)
            
        # Lire uniquement la feuille sélectionnée
        df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)

    logins_concernes = ["pvt_mwadk0290", "pvt_mwadk194", "pvt_mwadk181", "pvt_mwadk236",
        "pvt_sosy134", "pvt_sosy0290", "pvt_sosy0560", "pvt_sosy165",
        "pvt_dfallf0271", "pvt_dfallf0182", "pvt_dfallf0272", "pvt_dfallf0220",
        "Pvt_mbpling114", "Pvt_mbpling009", "Pvt_mbpling0230", "Pvt_mbpling173",
        "pvt_smmc301", "pvt_smmc2695", "pvt_smmc303", "pvt_smmc653",
        "pvt_tcg_0260", "pvt_tcg_0331", "pvt_tcg_0124", "pvt_tcg_0035"]

    details = ["En Cours-Identification", "Identifie", "Identifie Photo", ]
       
    # Nettoyage / Préparation
    df = df.rename(columns={'MSISDN': 'TOTAL_SIM'})
    df = df.rename(columns={'ACCUEIL_VENDEUR': 'PVT'})
    df = df.rename(columns={'LOGIN_VENDEUR': 'LOGIN'})
    df = df.rename(columns={'AGENCE_VENDEUR': 'DRV'})
        
    def clean_cols(df):
        df['DRV'] = df['DRV'].astype(str).str.strip().str.upper()
        #df['PVT'] = df['PVT'].astype(str).str.strip().str.upper()
        df['NOM_VENDEUR'] = df['NOM_VENDEUR'].astype(str).str.strip().str.upper()
        df['PRENOM_VENDEUR'] = df['PRENOM_VENDEUR'].astype(str).str.strip().str.upper()
        return df

    df = clean_cols(df)

    # 🔎 Filtrer les ventes LOUMA
    df_filtre = df[df['LOGIN'].astype(str).isin(logins_concernes) & df['ETAT_IDENTIFICATION'].astype(str).isin(details)]
    st.write("📊 Ventes LOUMA hebdomadaire :", df_filtre.shape[0], "lignes")

    st.success(f"✅ Feuille chargée avec succès !")
    st.dataframe(df.head())


    # -------- Résumé par VTO --------

    df_summary = df_filtre.groupby(['DRV', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN']).agg({
    'TOTAL_SIM': 'count'}).reset_index()  

    # Trier les données pour regrouper visuellement
    df_summary = df_summary.sort_values(['DRV', 'PVT'])


    df_summary["DRV"] = df_summary["DRV"].replace({ 
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DR SUD",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "SUD EST",
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DR NORD",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DR CENTRE",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DR EST"
        })


    # Pour masquer les répétitions (laisser vide sauf première occurrence)
    df_summaryy = df_summary.copy()
    df_summaryy['DRV'] = df_summaryy['DRV'].mask(df_summaryy['DRV'].duplicated())
    df_summaryy['PVT'] = df_summaryy['PVT'].mask(df_summaryy['PVT'].duplicated())


    # -------- Ventes par PVT (si dispo) --------
    df_summary2 = df_filtre.groupby(['DRV', 'PVT']).agg({
    'TOTAL_SIM': 'count'}).reset_index()
    df_summary2["OBJECTIF"] = 240
    df_summary2["TR"] = (df_summary2['TOTAL_SIM'] / df_summary2['OBJECTIF']).apply(lambda x: f"{round(x*100)}%")
    # Calculs : somme pour TOTAL_SIM et OBJECTIF, moyenne pour TR
    total_sim_sum = df_summary2['TOTAL_SIM'].sum()
    objectif_sum = df_summary2['OBJECTIF'].sum()
    tr_mean = df_summary2['TR'].apply(lambda x: float(x.strip('%'))).mean()

    # Ajout de la ligne "Total"
    df_summary2.loc['Total'] = [
        '',  # DRV
        '',  # PVT
        total_sim_sum,
        objectif_sum,
        f'{tr_mean:.1f}%'
    ]
 
    df_summary2["DRV"] = df_summary2["DRV"].replace({ 
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DR SUD",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "SUD EST",
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DR NORD",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DR CENTRE",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DR EST"
        })


    # 1. Créer un fichier Excel temporaire avec pandas
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        df_summaryy.to_excel(writer, sheet_name='Résumé Ventes', index=False)
        df_summary2.to_excel(writer, sheet_name='Ventes Par PVT', index=False)

    # 2. Charger avec openpyxl pour appliquer la fusion
    wb = load_workbook(temp_file.name)
    ws = wb["Résumé Ventes"]



    # 3. Sauvegarde dans un buffer pour Streamlit
    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
        

        #----------------------------------------------------------------------------------------------------

    # Télécharger le fichier généré
    st.success("✅ Fichier généré avec succès !")
    st.download_button(
        label="📥 Télécharger le fichier Excel",
        data=final_buffer,
        file_name="Weekly Reporting.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




    import pandas as pd
    from datetime import datetime
    import re
    from io import BytesIO
    import os

    fichier_historique = r"C:\Users\hp\Downloads\Dossier LOUMA\historique_ventes_.xlsx"

    # 1. Préparer df_summary
    df_history = df_summary.copy()

    # 2. Identifier la dernière semaine existante
    if os.path.exists(fichier_historique):
        df_old = pd.read_excel(fichier_historique)

        # Extraire toutes les colonnes de type SEMxx
        sem_cols = [col for col in df_old.columns if re.match(r"SEM\d+", str(col))]

        if sem_cols:
            derniers_num_semaines = [int(col[3:]) for col in sem_cols]
            max_semaine = max(derniers_num_semaines)
            nouvelle_semaine = max_semaine + 1
        else:
            nouvelle_semaine = 1

        nom_col_semaine = f"SEM{nouvelle_semaine}"
        st.info(f"🕓 Semaine détectée : {nom_col_semaine}")

        #
        #yallah pitié
        def create_key(df):
            return (
            df_summary['DRV'].astype(str).str.strip().str.upper() + "|" +
            df_summary['PVT'].astype(str).str.strip().str.upper() + "|" +
            df_summary['PRENOM_VENDEUR'].astype(str).str.strip().str.upper() + "|" +
            df_summary['NOM_VENDEUR'].astype(str).str.strip().str.upper()
        )

        
        #
        # Renommer la colonne dans df_summary
        df_history = df_history.rename(columns={'TOTAL_SIM': nom_col_semaine})
        #
        df_history_clean = df_history[['DRV', 'PRENOM_VENDEUR', 'NOM_VENDEUR', nom_col_semaine]]
        
        ##
        # Ajouter la colonne au fichier historique ligne par ligne
        #if nom_col_semaine not in df_old.columns:
            # df_old[nom_col_semaine] = 0

        # Nettoyer les champs d'identification dans les deux DataFrames
        

        df_old = clean_cols(df_old)
        df_history = clean_cols(df_history)

        
        n_mis_a_jour = 0
        n_non_trouves = 0
        ##
        # 5. Appliquer les mises à jour à partir de df_summary
        for idx_old, row_old in df_old.iterrows():
            match = df_summary[
                (df_summary['DRV'] == row_old['DRV']) &
                (df_summary['PRENOM_VENDEUR'] == row_old['PRENOM_VENDEUR']) &
                (df_summary['NOM_VENDEUR'] == row_old['NOM_VENDEUR'])
            ]
            if not match.empty:
                total = match.iloc[0]['TOTAL_SIM']
                df_old.at[idx_old, nom_col_semaine] = total
                n_mis_a_jour += 1
                st.write(f"✅ {row_old['PRENOM_VENDEUR']} {row_old['NOM_VENDEUR']} mis à jour avec {total} dans {nom_col_semaine}")
            else:
                n_non_trouves += 1
                st.warning(f"❌ {row_old['PRENOM_VENDEUR']} {row_old['NOM_VENDEUR']} introuvable dans df_summary")
                df_old.at[idx_old, nom_col_semaine] = 0
        # Fusionner avec l'existant
        #df_merged = pd.merge(df_old, df_history_clean, on=['DRV', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR'], how='outer')


    #else:
        #nom_col_semaine = "SEM1"
        #st.info(f"📁 Fichier historique non trouvé : création avec {nom_col_semaine}")
        #df_history = df_history.rename(columns={'TOTAL_SIM': nom_col_semaine})
        #df_merged = df_history.copy()

    # 3. Sauvegarder le fichier mis à jour
    df_old.to_excel(fichier_historique, index=False)

    st.success(f"✅ Données ajoutées à l’historique (colonne : {nom_col_semaine})")

    # 4. Ajouter un bouton pour télécharger le fichier historique
    with open(fichier_historique, "rb") as file:
        st.download_button(
            label="📥 Télécharger le fichier historique",
            data=file,
            file_name="historique_ventes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
